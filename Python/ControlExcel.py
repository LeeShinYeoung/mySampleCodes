import xlrd
import xlwt
import re
import sys
import pprint
import openpyxl
import numpy as np
from openpyxl import Workbook
from pathlib import Path
from xlutils.copy import copy

# ==== ControlExcel ====
# 엑셀 컨트롤 클래스 ( .xls, .xlsx 가능 )
# path = 엑셀파일경로
# inst = ControlExcel(path)
#
# ==== exec_ 동작 ====
# 셀 반환
# exec_get_cell(row, col)
#
# 행렬 반환
# exec_get_matrix(self)
#
# 행렬 삽입
# exec_append_matrix(matrix)
#
# self.path에 paths에 있는 데이터 모두 삽입
# exec_append_paths(paths):
#
# self.path에 파일 생성
# exec_make_file(path)
#
# ==== set_ 동작전 데이터 설정 (필수x) ====
# 시트 설정 ( exec_ 실행전 실행 )
# set_sheet(sheet)
#
# 가져올 컬럼 설정 ( exec_ 실행전 실행 )
# set_columns(cols_list):
#
# ==== get_ 조회 ====
# 컬럼 필터링
# get_filtred_column(matrix):
#
# 행, 열, 시트 갯수 반환
# get_count(type)
#
# self.print_process => True면 실행순서 출력
# show_process(data)
#
# 설정 경로에 대해 같은 종류의 파일 반환
# @classmethod
# get_same_file_paths(dir_path, file_name, file_format)
#
# 파일 유무 반환
# @classmethod
# is_file(cls, path)


class ControlExcel:
    def __init__(self, path=False, print_process=False):
        if not self.is_file(path):
            self.exec_make_file(path)
        self.print_process = print_process if print_process else False
        self.path = path
        self.file_format = re.findall('\.xls$|\.xlsx$', self.path)[0]
        self.workbench = xlrd.open_workbook(path)
        self.worksheet = False
        self.sheet = 0
        self.appointed_cols = False

    # 셀 반환
    def exec_get_cell(self, row, col):
        result = self.worksheet.cell_value(row, col)
        self.show_process(result)
        return result

    # 행렬 반환
    def exec_get_matrix(self):
        if not self.worksheet:
            self.set_sheet()
        cnt = self.get_count('rows')
        filtered_matrix = False
        matrix = []
        # 데이터 입력
        for i in range(cnt):
            matrix.append(self.worksheet.row_values(i))
        # 컬럼 필터링
        if self.appointed_cols:
            filtered_matrix = self.get_filtred_column(matrix)
        result = filtered_matrix if filtered_matrix else matrix
        self.show_process(result)
        return matrix

    # 행렬 삽입
    def exec_append_matrix(self, matrix):
        try:
            # xls 저장
            if self.file_format == '.xls':
                if not self.worksheet:
                    self.set_sheet()
                start = self.get_count('rows')
                workbench_cp = copy(self.workbench)
                # 시트 선택
                if isinstance(self.sheet, int):
                    worksheet_cp = workbench_cp.get_sheet(self.sheet)
                if isinstance(self.sheet, str):
                    worksheet_cp = workbench_cp.get_sheet(self.sheet)
                # 데이터 입력
                for i, row in enumerate(matrix):
                    for j, value in enumerate(row):
                        worksheet_cp.write(start + i, j, value)
                workbench_cp.save(self.path)

            # xlsx 저장
            if self.file_format == '.xlsx':
                wb = openpyxl.load_workbook(self.path)
                sheets = wb.get_sheet_names()
                # 시트 선택
                if self.sheet in sheets:
                    worksheet = wb.get_sheet_by_name(self.sheet)
                if isinstance(self.sheet, int):
                    worksheet = wb.worksheets[self.sheet]
                # 데이터 입력
                start = worksheet.max_row
                for i, row in enumerate(matrix, 1):
                    for j, value in enumerate(row, 1):
                        worksheet.cell(start + i, j, value)
                wb.save(self.path)
        except PermissionError:
            self.show_process('Error : 권한이 없습니다. 실행중인 파일을 닫아주세요.')
            return False

        self.show_process("append {} rows".format(len(matrix)))
        return True

    # self.path에 paths에 있는 데이터 모두 삽입
    def exec_append_paths(self, paths):
        # 파일 유무 체크
        for path in paths:
            if not self.is_file(path):
                self.show_process('Error : {}에 파일이 없습니다'.format(path))
                return False
        # paths에 있는 파일들 self.path에 모두 삽입
        for path in paths:
            instance = ControlExcel(path)
            data = instance.exec_get_matrix()
            self.exec_append_matrix(data)
            self.show_process(path)

    # self.path에 파일 생성
    def exec_make_file(self, path=False):
        path = path if path else self.path
        file_format = re.findall('\.xls$|\.xlsx$', path)[0]
        # xls 파일 생성
        if file_format == '.xls':
            workbench = xlwt.Workbook(encoding='utf-8')
            workbench.add_sheet('sheet1')
            workbench.save(path)
        # xlsx 파일 생성
        if file_format == '.xlsx':
            workbench = Workbook()
            workbench.create_sheet('sheet1')
            workbench.save(path)

    # 시트 설정 ( exec_ 실행전 실행 )
    def set_sheet(self, sheet=None):
        self.sheet = sheet if sheet else self.sheet
        if isinstance(self.sheet, int):
            self.worksheet = self.workbench.sheet_by_index(self.sheet)
        if isinstance(self.sheet, str):
            self.worksheet = self.workbench.sheet_by_name(self.sheet)
        self.show_process("Select Sheet : {}".format(self.sheet))

    # 가져올 컬럼 설정 ( exec_ 실행전 실행 )
    def set_columns(self, cols_list):
        self.appointed_cols = cols_list
        self.show_process("Select Columns : {}".format(self.appointed_cols))
        return self.appointed_cols

    # 컬럼 필터링
    def get_filtred_column(self, matrix):
        # 열 필터링과정
        # 전치행렬 -> 행 필터링 -> 전치행렬
        filterd_matrix = np.transpose([
            cols.tolist()
            for cols in np.transpose(matrix)
            for cell in cols
            if cell in self.appointed_cols]).tolist()
        self.show_process(filterd_matrix)
        return filterd_matrix

    # 행, 열, 시트 갯수 반환
    def get_count(self, type):
        if not type:
            return False
        if not self.worksheet:
            return False
        elif type in ['col', 'cols', 'column', 'columns']:
            count = self.worksheet.ncols
        elif type in ['row', 'rows']:
            count = self.worksheet.nrows
        elif type in ['sheet', 'sheets']:
            count = self.worksheet.nsheets
        else:
            return False
        self.show_process(count)
        return count

    # self.print_process => True면 실행순서 출력
    def show_process(self, data):
        if not self.print_process:
            return False
        method_name = sys._getframe().f_back.f_code.co_name
        print('** ' + method_name + ' **')
        if isinstance(data, list):
            pprint.pprint(data)
        else:
            print(data)
        return data

    # 설정 경로에 대해 같은 종류의 파일 반환
    @classmethod
    def get_same_file_paths(cls, dir_path, file_name, file_format):
        i = 0
        paths = []
        while True:
            f = file_name + ' (' + str(i) + ')' if i != 0 else file_name
            path = dir_path + f + file_format
            if not cls.is_file(path):
                return str(len(paths)) + '개 파일 발견\n' + str(paths)
            paths.append(path)
            i = i + 1
        return paths

    # 파일 유무 반환
    @classmethod
    def is_file(cls, path):
        check = Path(path).is_file()
        return check


if __name__ == '__main__':
    pass
